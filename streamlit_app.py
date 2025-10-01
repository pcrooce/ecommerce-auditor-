import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import time
import requests
from bs4 import BeautifulSoup
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    st.warning("Playwright no instalado. Frávega tendrá funcionalidad limitada.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("Instala: pip install openpyxl")

st.set_page_config(
    page_title="Auditor Automático",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
.main {padding: 0rem 1rem;}
.audit-header {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    padding: 2.5rem; border-radius: 15px; color: white;
    margin-bottom: 2rem; box-shadow: 0 10px 30px rgba(0,0,0,0.2);
    text-align: center;
}
div[data-testid="metric-container"] {
    background-color: #f8f9fa; border: 2px solid #e9ecef;
    padding: 15px; border-radius: 10px; margin: 10px 0px;
}
.stButton > button {
    background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    color: white; border: none; padding: 0.5rem 1rem;
    font-weight: 600; border-radius: 8px;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="audit-header">
    <h1>🤖 Sistema de Auditoría Automática v6.0</h1>
    <p>Con soporte completo para Frávega</p>
</div>
""", unsafe_allow_html=True)

if 'audit_results' not in st.session_state:
    st.session_state.audit_results = None

TIENDAS_CONFIG = {
    "ICBC": {
        "columnas_busqueda": ["ICBC", "icbc"],
        "selector_titulo": "h1[itemprop='name']",
        "selector_precio": "p.monto",
        "selector_precio_tachado": "p.precio-anterior",
        "selector_descuento": "p.descuento",
        "selector_categoria": "span.breadcrumb-span[itemprop='title']"
    },
    "Supervielle": {
        "columnas_busqueda": ["Supervielle", "supervielle"],
        "selector_titulo": "h1[itemprop='name']",
        "selector_precio": "span#our_price_display",
        "selector_precio_tachado": "span.price",
        "selector_descuento": "span#reduction_percent_display",
        "selector_categoria": "span[itemprop='title']"
    },
    "Galicia": {
        "columnas_busqueda": ["Galicia", "galicia"],
        "selector_titulo": "h1.productTitle",
        "selector_precio": "div.productPrice span",
        "selector_descuento": "span.discount.discount-percentage",
        "selector_categoria": "span[itemprop='name']"
    },
    "Ciudad": {
        "columnas_busqueda": ["Ciudad", "ciudad"],
        "selector_titulo": "h1.name",
        "selector_precio": "span.amount",
        "selector_precio_tachado": "div[itemprop='offers'] span.amount",
        "selector_categoria": "a[href*='/catalog/']"
    },
    "Fravega": {
        "columnas_busqueda": ["Fravega", "fravega", "FVG", "fvg"],
        "columnas_cuotas": ["Cuotas FVG", "CSI FVG", "Financiacion Fvg", "Financiación FVG", "cuotas fvg", "csi fvg"],
        "selector_titulo": "h1[data-test-id='product-title']",
        "selector_precio": "span.sc-1d9b1d9e-0.sc-faa1a185-3",
        "selector_precio_tachado": "span.sc-e081bce1-0.sc-faa1a185-4",
        "selector_descuento": "span.sc-e2aca368-0",
        "selector_categoria": "span[itemprop='name']",
        "selector_cuotas_container": "span.sc-3cba7521-10",
        "urls_visa_master": ["54c0d769ece1b", "d91d7904a8578", "visa", "mastercard"]
    },
    "BNA": {
        "columnas_busqueda": ["BNA", "bna"],
        "selector_precio": "span.price"
    },
    "Megatone": {
        "columnas_busqueda": ["Megatone", "megatone", "MGT", "mgt"],
        "columnas_cuotas": ["Cuotas MGT", "CSI MGT", "cuotas mgt", "csi mgt"],
        "selector_precio": "span.price"
    }
}

def detectar_columnas_automaticamente(df, tienda):
    config = TIENDAS_CONFIG[tienda]
    resultado = {'url': None, 'precio': None, 'sku': None, 'cuotas': None}
    
    # Patrones más amplios para detección
    patrones_url = []
    patrones_precio = []
    patrones_cuotas = []
    
    for busqueda in config['columnas_busqueda']:
        patrones_url.extend([f'{busqueda} url', f'url {busqueda}', f'{busqueda} link', f'link {busqueda}'])
        patrones_precio.extend([f'pvp {busqueda}', f'{busqueda} pvp', f'precio {busqueda}', f'{busqueda} precio'])
    
    if 'columnas_cuotas' in config:
        for busqueda in config['columnas_cuotas']:
            patrones_cuotas.append(busqueda.lower())
    
    for col in df.columns:
        col_lower = col.lower().strip()
        
        if resultado['url'] is None:
            for patron in patrones_url:
                if patron.lower() in col_lower:
                    resultado['url'] = col
                    break
        
        if resultado['precio'] is None:
            for patron in patrones_precio:
                if patron.lower() in col_lower:
                    resultado['precio'] = col
                    break
        
        if resultado['sku'] is None:
            if any(word in col_lower for word in ['sku', 'codigo', 'código']):
                resultado['sku'] = col
        
        if 'columnas_cuotas' in config and resultado['cuotas'] is None:
            for patron in patrones_cuotas:
                if patron in col_lower:
                    resultado['cuotas'] = col
                    break
    
    return resultado

def limpiar_precio(valor):
    if pd.isna(valor):
        return None
    
    precio_str = str(valor).replace('$', '').replace(' ', '').strip()
    
    if not precio_str:
        return None
    
    if '.' in precio_str and ',' in precio_str:
        precio_str = precio_str.replace('.', '').replace(',', '.')
    elif '.' in precio_str:
        if re.search(r'\.\d{3}', precio_str):
            precio_str = precio_str.replace('.', '')
        elif not re.search(r'\.\d{2}$', precio_str):
            precio_str = precio_str.replace('.', '')
    elif ',' in precio_str:
        if re.search(r',\d{2}$', precio_str):
            precio_str = precio_str.replace(',', '.')
        else:
            precio_str = precio_str.replace(',', '')
    
    try:
        valor_float = float(re.sub(r'[^\d.]', '', precio_str))
        return valor_float if valor_float > 0 else None
    except:
        return None

class WebScraper:
    def __init__(self, tienda_config, tienda_nombre):
        self.config = tienda_config
        self.tienda = tienda_nombre
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'es-AR,es;q=0.9,en;q=0.8',
            'Referer': 'https://www.google.com/'
        })
    
    def scrape_fravega_con_playwright(self, url):
        """Scrapea Frávega usando Playwright para contenido dinámico"""
        
        # VALIDACIÓN: URL debe ser válida
        if not url or not isinstance(url, str):
            return {
                'url': url,
                'titulo': None,
                'precio_web': None,
                'precio_tachado': None,
                'descuento_%': None,
                'categoria': None,
                'cuotas': None,
                'estado_producto': 'Error',
                'estado_scraping': '❌ URL inválida',
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        
        # VALIDACIÓN: URL debe comenzar con http:// o https://
        if not url.startswith('http://') and not url.startswith('https://'):
            return {
                'url': url,
                'titulo': None,
                'precio_web': None,
                'precio_tachado': None,
                'descuento_%': None,
                'categoria': None,
                'cuotas': None,
                'estado_producto': 'Error',
                'estado_scraping': '❌ URL incompleta - falta https://',
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        
        # VALIDACIÓN: URL muy corta
        if len(url) < 30:
            return {
                'url': url,
                'titulo': None,
                'precio_web': None,
                'precio_tachado': None,
                'descuento_%': None,
                'categoria': None,
                'cuotas': None,
                'estado_producto': 'Error',
                'estado_scraping': '❌ URL demasiado corta',
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        
        resultado = {
            'url': url,
            'titulo': None,
            'precio_web': None,
            'precio_tachado': None,
            'descuento_%': None,
            'categoria': None,
            'cuotas': None,
            'estado_producto': 'Activo',
            'estado_scraping': '✅ OK',
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context(
                    viewport={'width': 1920, 'height': 1080},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                )
                page = context.new_page()
                
                page.goto(url, wait_until='networkidle', timeout=30000)
                page.wait_for_timeout(3000)
                
                # PRIMERO: Verificar si está inhabilitado
                producto_inhabilitado = False
                try:
                    boton = page.locator("button[data-test-id='product-buy-button']").first
                    boton.wait_for(timeout=5000)
                    
                    is_disabled = boton.is_disabled()
                    has_disabled_attr = boton.get_attribute('disabled') is not None
                    
                    if is_disabled or has_disabled_attr:
                        producto_inhabilitado = True
                    
                    texto_boton = boton.text_content()
                    if texto_boton and 'no disponible' in texto_boton.lower():
                        producto_inhabilitado = True
                        
                except:
                    producto_inhabilitado = True
                
                # Título (siempre intentar obtener)
                try:
                    titulo = page.locator("h1[data-test-id='product-title']").text_content(timeout=5000)
                    if titulo:
                        resultado['titulo'] = titulo.strip()
                except:
                    pass
                
                # CORRECCIÓN: Categorías - excluir "Frávega" y tomar última válida
                try:
                    categorias_elems = page.locator("span[itemprop='name']").all()
                    categorias_validas = []
                    for elem in categorias_elems:
                        texto = elem.text_content().strip()
                        # Excluir explícitamente nombres de tiendas
                        if texto and texto.lower() not in ['frávega', 'fravega', 'inicio', 'home']:
                            categorias_validas.append(texto)
                    
                    # Tomar la ÚLTIMA categoría válida
                    if categorias_validas:
                        resultado['categoria'] = categorias_validas[-1]
                except:
                    pass
                
                # Si está inhabilitado, marcar correctamente y NO scrapear precios
                if producto_inhabilitado:
                    resultado['estado_producto'] = 'Inhabilitado'
                    resultado['estado_scraping'] = '⚠️ Botón de compra deshabilitado'
                    resultado['cuotas'] = None
                    browser.close()
                    return resultado
                
                # Precio (solo si está habilitado)
                try:
                    precio = page.locator("span.sc-1d9b1d9e-0.sc-faa1a185-3").first.text_content(timeout=5000)
                    resultado['precio_web'] = limpiar_precio(precio)
                except:
                    pass
                
                # Precio tachado
                try:
                    tachado = page.locator("span.sc-e081bce1-0.sc-faa1a185-4").first.text_content(timeout=5000)
                    resultado['precio_tachado'] = limpiar_precio(tachado)
                except:
                    pass
                
                # Descuento
                try:
                    descuento = page.locator("span.sc-e2aca368-0").first.text_content(timeout=5000)
                    match = re.search(r'(\d+)', descuento)
                    if match:
                        resultado['descuento_%'] = float(match.group(1))
                except:
                    pass
                
                # CORRECCIÓN CRÍTICA: Cuotas - SOLO primeras 2 imágenes (Visa y Mastercard)
                cuotas_encontradas = False
                try:
                    page_content = page.content()
                    soup = BeautifulSoup(page_content, 'html.parser')
                    
                    cuotas_divs = soup.find_all('div', class_=lambda x: x and 'sc-3cba7521-0' in x)
                    
                    for div in cuotas_divs:
                        cuotas_span = div.find('span', class_=lambda x: x and 'sc-3cba7521-10' in x)
                        
                        if not cuotas_span:
                            continue
                        
                        texto = cuotas_span.get_text()
                        match = re.search(r'(\d+)\s*cuotas?', texto, re.IGNORECASE)
                        
                        if not match:
                            continue
                        
                        num_cuotas = int(match.group(1))
                        
                        img_container = div.find('div', class_=lambda x: x and 'sc-3cba7521-3' in x)
                        
                        if img_container:
                            imagenes = img_container.find_all('img', src=True)
                            
                            # CRÍTICO: Solo verificar las primeras 2 imágenes
                            if len(imagenes) >= 2:
                                img1_src = imagenes[0].get('src', '').lower()
                                img2_src = imagenes[1].get('src', '').lower()
                                
                                # Verificar que las primeras 2 sean Visa o Mastercard
                                es_visa_master = ('d91d7904a8578' in img1_src or '54c0d769ece1b' in img1_src or
                                                'd91d7904a8578' in img2_src or '54c0d769ece1b' in img2_src)
                                
                                if es_visa_master:
                                    resultado['cuotas'] = num_cuotas
                                    cuotas_encontradas = True
                                    break
                    
                    if not cuotas_encontradas:
                        resultado['cuotas'] = 1
                        
                except Exception as e:
                    resultado['cuotas'] = 1
                    resultado['estado_scraping'] = f'⚠️ OK (error cuotas: {str(e)[:20]})'
                
                # Validar que se haya scrapeado el precio
                if not resultado['precio_web']:
                    resultado['estado_scraping'] = '⚠️ No se obtuvo el precio'
                
                browser.close()
                
        except Exception as e:
            resultado['estado_producto'] = 'Error'
            resultado['estado_scraping'] = f'❌ {str(e)[:40]}'
        
        return resultado
    
    def scrape_url(self, url):
        # CAMBIO CRÍTICO: Si es Frávega, usar Playwright directamente
        if self.tienda == "Fravega":
            if PLAYWRIGHT_AVAILABLE:
                return self.scrape_fravega_con_playwright(url)
            else:
                return {
                    'url': url,
                    'titulo': None,
                    'precio_web': None,
                    'precio_tachado': None,
                    'descuento_%': None,
                    'categoria': None,
                    'cuotas': None,
                    'estado_producto': 'Error',
                    'estado_scraping': '❌ Playwright no disponible',
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
        
        # Para otras tiendas, usar requests
        resultado = {
            'url': url,
            'titulo': None,
            'precio_web': None,
            'precio_tachado': None,
            'descuento_%': None,
            'categoria': None,
            'cuotas': None,
            'estado_producto': 'Activo',
            'estado_scraping': '✅ OK',
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        try:
            response = self.session.get(url, timeout=15)
            
            if response.status_code == 404:
                resultado['estado_producto'] = 'No disponible'
                resultado['estado_scraping'] = '⚠️ Error 404'
                return resultado
            
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            
            html_text = soup.get_text().lower()
            if 'no longer available' in html_text or 'no está disponible' in html_text:
                resultado['estado_producto'] = 'No disponible'
                resultado['estado_scraping'] = '⚠️ Producto no disponible'
                return resultado
            
            if 'selector_precio' in self.config:
                precio_elem = soup.select_one(self.config['selector_precio'])
                if precio_elem:
                    resultado['precio_web'] = limpiar_precio(precio_elem.get_text(strip=True))
            
            if 'selector_precio_tachado' in self.config:
                tachado_elem = soup.select_one(self.config['selector_precio_tachado'])
                if tachado_elem:
                    resultado['precio_tachado'] = limpiar_precio(tachado_elem.get_text(strip=True))
            
            if 'selector_descuento' in self.config:
                desc_elem = soup.select_one(self.config['selector_descuento'])
                if desc_elem:
                    desc_text = desc_elem.get_text(strip=True)
                    match = re.search(r'(\d+)', desc_text)
                    if match:
                        resultado['descuento_%'] = float(match.group(1))
            
            if self.tienda == "Galicia" and not resultado['precio_tachado'] and resultado['descuento_%'] and resultado['precio_web']:
                descuento_decimal = resultado['descuento_%'] / 100
                resultado['precio_tachado'] = resultado['precio_web'] / (1 - descuento_decimal)
            
            if not resultado['precio_web']:
                resultado['estado_scraping'] = '⚠️ No se obtuvo el precio'
            
        except requests.exceptions.HTTPError as e:
            if '404' in str(e):
                resultado['estado_producto'] = 'No disponible'
                resultado['estado_scraping'] = '⚠️ Error 404'
            else:
                resultado['estado_producto'] = 'Error'
                resultado['estado_scraping'] = f'❌ {str(e)[:30]}'
        except Exception as e:
            resultado['estado_producto'] = 'Error'
            resultado['estado_scraping'] = f'❌ {str(e)[:30]}'
        
        return resultado

def realizar_scraping(df_tienda, tienda_config, tienda_nombre, progress_bar, status_text):
    scraper = WebScraper(tienda_config, tienda_nombre)
    resultados = []
    
    # Para Frávega, hacer scraping secuencial (Playwright no es thread-safe)
    if tienda_nombre == "Fravega":
        for idx, row in df_tienda.iterrows():
            if pd.notna(row.get('url')):
                resultado = scraper.scrape_url(row['url'])
                resultado['idx'] = idx
                resultados.append(resultado)
                
                completed = idx + 1
                total = len(df_tienda)
                progress_bar.progress(min(completed / total, 1.0))
                status_text.text(f"Escaneando {completed}/{total}...")
    else:
        # Para otras tiendas, usar ThreadPool
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(scraper.scrape_url, row['url']): idx 
                      for idx, row in df_tienda.iterrows() if pd.notna(row.get('url'))}
            
            completed = 0
            total = len(futures)
            
            for future in as_completed(futures):
                completed += 1
                idx = futures[future]
                resultado = future.result()
                resultado['idx'] = idx
                resultados.append(resultado)
                
                progress_bar.progress(min(completed / total, 1.0))
                status_text.text(f"Escaneando {completed}/{total}...")
    
    return resultados

def crear_excel_formateado(df_results, tienda):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    
    ws['A1'] = f'AUDITORÍA {tienda.upper()} - {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A1'].font = Font(bold=True, size=14)
    
    if tienda in ["Fravega", "Megatone"]:
        columnas = ['SKU', 'Título', 'Precio Maestro', 'Precio Web', 'Precio Tachado',
                   'Descuento %', 'Variación %', 'Precio OK', 'Cuotas Maestro', 'Cuotas Web',
                   'Cuotas OK', 'Categoría', 'Estado', 'Scraping', 'URL']
        ws.merge_cells('A1:O1')
    else:
        columnas = ['SKU', 'Título', 'Precio Maestro', 'Precio Web', 'Precio Tachado',
                   'Descuento %', 'Variación %', 'Precio OK', 'Categoría', 'Estado', 
                   'Scraping', 'URL']
        ws.merge_cells('A1:L1')
    
    ws.append([])
    ws.append(columnas)
    
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    
    for _, row in df_results.iterrows():
        if tienda in ["Fravega", "Megatone"]:
            row_data = [
                row.get('sku'), row.get('titulo'), row.get('precio_maestro'),
                row.get('precio_web'), row.get('precio_tachado'), row.get('descuento_%'),
                row.get('variacion_precio_%'),
                'Sí' if row.get('precio_ok') == True else 'No' if row.get('precio_ok') == False else '-',
                row.get('cuotas_maestro'), row.get('cuotas'),
                'Sí' if row.get('cuotas_correctas') == True else 'No' if row.get('cuotas_correctas') == False else '-',
                row.get('categoria'), row.get('estado_producto'), 
                row.get('estado_scraping'), row.get('url')
            ]
        else:
            row_data = [
                row.get('sku'), row.get('titulo'), row.get('precio_maestro'),
                row.get('precio_web'), row.get('precio_tachado'), row.get('descuento_%'),
                row.get('variacion_precio_%'),
                'Sí' if row.get('precio_ok') == True else 'No' if row.get('precio_ok') == False else '-',
                row.get('categoria'), row.get('estado_producto'),
                row.get('estado_scraping'), row.get('url')
            ]
        ws.append(row_data)
    
    # Ajustar ancho de columnas
    for idx in range(1, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 15
    
    wb.save(output)
    output.seek(0)
    return output

with st.sidebar:
    st.markdown("""
        <div style='text-align: center; padding: 1rem; 
             background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
             border-radius: 10px; margin-bottom: 1rem;'>
            <h2 style='color: white; margin: 0;'>⚙️ Panel</h2>
        </div>
    """, unsafe_allow_html=True)
    
    selected_store = st.selectbox("🏪 Tienda", list(TIENDAS_CONFIG.keys()))
    
    if selected_store == "Fravega":
        if not PLAYWRIGHT_AVAILABLE:
            st.error("⚠️ Playwright requerido para Frávega")
            st.info("Ejecutar: pip install playwright && playwright install chromium")
        else:
            st.success("✅ Playwright listo")
    
    price_threshold = st.slider("🎯 Tolerancia (%)", 0, 20, 5, 1)
    
    modo_operacion = st.radio("🚀 Modo", [
        "🧪 Prueba (simulado)",
        "⚡ Rápida (10 productos)", 
        "📊 Completa"
    ])
    
    if "Prueba" in modo_operacion:
        modo_operacion = "Modo Prueba"
        max_productos = 10
    elif "Rápida" in modo_operacion:
        modo_operacion = "Auditoría Rápida"
        max_productos = 10
    else:
        modo_operacion = "Auditoría Completa"
        max_productos = st.number_input("Límite:", 10, 1000, 100, 10)

tab1, tab2, tab3 = st.tabs(["📁 Cargar", "📊 Resultados", "📈 Dashboard"])

with tab1:
    st.markdown("### 📝 Auditoría")
    
    uploaded_file = st.file_uploader("Cargar Excel", type=['xlsx', 'xls'])
    
    if uploaded_file:
        df_maestro = pd.read_excel(uploaded_file)
        
        col1, col2, col3 = st.columns(3)
        col1.metric("📄 Archivo", uploaded_file.name[:15] + "...")
        col2.metric("📊 Filas", f"{len(df_maestro):,}")
        col3.metric("📋 Columnas", len(df_maestro.columns))
        
        columnas_detectadas = detectar_columnas_automaticamente(df_maestro, selected_store)
        
        todas_detectadas = all([columnas_detectadas['url'], columnas_detectadas['sku'], columnas_detectadas['precio']])
        if selected_store in ["Fravega", "Megatone"]:
            todas_detectadas = todas_detectadas and columnas_detectadas['cuotas']
        
        if todas_detectadas:
            st.success("✅ Todas las columnas detectadas automáticamente")
            col1, col2 = st.columns(2)
            col1.info(f"📍 URL: **{columnas_detectadas['url']}**")
            col1.info(f"🏷️ SKU: **{columnas_detectadas['sku']}**")
            col2.info(f"💰 Precio: **{columnas_detectadas['precio']}**")
            if selected_store in ["Fravega", "Megatone"]:
                col2.info(f"💳 Cuotas: **{columnas_detectadas['cuotas']}**")
            
            url_column = columnas_detectadas['url']
            sku_column = columnas_detectadas['sku']
            precio_column = columnas_detectadas['precio']
            cuotas_column = columnas_detectadas['cuotas'] if selected_store in ["Fravega", "Megatone"] else None
        else:
            st.warning("⚠️ Seleccione columnas manualmente:")
            
            col1, col2 = st.columns(2)
            
            with col1:
                url_column = st.selectbox("URL:", df_maestro.columns, 
                                         index=list(df_maestro.columns).index(columnas_detectadas['url']) if columnas_detectadas['url'] else 0)
                sku_column = st.selectbox("SKU:", df_maestro.columns,
                                         index=list(df_maestro.columns).index(columnas_detectadas['sku']) if columnas_detectadas['sku'] else 0)
            
            with col2:
                precio_column = st.selectbox("Precio:", df_maestro.columns,
                                            index=list(df_maestro.columns).index(columnas_detectadas['precio']) if columnas_detectadas['precio'] else 0)
                
                if selected_store in ["Fravega", "Megatone"]:
                    cuotas_column = st.selectbox("Cuotas:", df_maestro.columns,
                                                index=list(df_maestro.columns).index(columnas_detectadas['cuotas']) if columnas_detectadas['cuotas'] else 0)
                else:
                    cuotas_column = None
        
        df_tienda = df_maestro[df_maestro[url_column].notna()].copy()
        
        rename_dict = {url_column: 'url', sku_column: 'sku', precio_column: 'precio_maestro'}
        if cuotas_column:
            rename_dict[cuotas_column] = 'cuotas_maestro'
        
        df_tienda = df_tienda.rename(columns=rename_dict)
        df_tienda['precio_maestro'] = df_tienda['precio_maestro'].apply(limpiar_precio)
        
        if 'cuotas_maestro' in df_tienda.columns:
            df_tienda['cuotas_maestro'] = pd.to_numeric(df_tienda['cuotas_maestro'], errors='coerce')
        
        df_tienda = df_tienda.head(max_productos)
        
        st.markdown("---")
        
        if st.button("🚀 INICIAR", type="primary", use_container_width=True):
            
            if "Prueba" in modo_operacion:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                resultados = []
                for i, (idx, row) in enumerate(df_tienda.iterrows()):
                    variacion = np.random.uniform(-10, 10)
                    precio_maestro_val = row.get('precio_maestro', 10000)
                    if pd.isna(precio_maestro_val):
                        precio_maestro_val = 10000
                    
                    precio_web = float(precio_maestro_val * (1 + variacion/100))
                    
                    resultados.append({
                        'idx': idx,
                        'url': row['url'],
                        'titulo': f"Producto Ejemplo {i+1}",
                        'precio_web': precio_web,
                        'precio_tachado': precio_web * 1.3,
                        'descuento_%': float(np.random.randint(10, 40)),
                        'categoria': "Categoría Ejemplo",
                        'cuotas': int(np.random.choice([1, 3, 6, 9, 12])) if selected_store in ["Fravega", "Megatone"] else None,
                        'estado_producto': 'Activo',
                        'estado_scraping': '✅ OK'
                    })
                    
                    progress_bar.progress((i + 1) / len(df_tienda))
                    status_text.text(f"{i + 1}/{len(df_tienda)}")
                    time.sleep(0.05)
                
                progress_bar.empty()
                status_text.empty()
            else:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                resultados = realizar_scraping(
                    df_tienda, 
                    TIENDAS_CONFIG[selected_store], 
                    selected_store, 
                    progress_bar, 
                    status_text
                )
                
                progress_bar.empty()
                status_text.empty()
            
            for resultado in resultados:
                idx = resultado['idx']
                df_tienda.loc[idx, 'titulo'] = resultado.get('titulo')
                df_tienda.loc[idx, 'precio_web'] = resultado.get('precio_web')
                df_tienda.loc[idx, 'precio_tachado'] = resultado.get('precio_tachado')
                df_tienda.loc[idx, 'descuento_%'] = resultado.get('descuento_%')
                df_tienda.loc[idx, 'categoria'] = resultado.get('categoria')
                df_tienda.loc[idx, 'cuotas'] = resultado.get('cuotas')
                df_tienda.loc[idx, 'estado_producto'] = resultado.get('estado_producto')
                df_tienda.loc[idx, 'estado_scraping'] = resultado.get('estado_scraping')
            
            # Calcular variación solo para activos con precio
            mask = ((df_tienda['precio_web'].notna()) & (df_tienda['precio_maestro'].notna()) & 
                    (df_tienda['precio_maestro'] > 0) & (df_tienda['estado_producto'] == 'Activo'))
            
            df_tienda['variacion_precio_%'] = None
            if mask.any():
                df_tienda.loc[mask, 'variacion_precio_%'] = ((df_tienda.loc[mask, 'precio_web'] - df_tienda.loc[mask, 'precio_maestro']) / 
                                                               df_tienda.loc[mask, 'precio_maestro'] * 100).round(2)
            
            # Precio OK solo si hay precio Y está en rango
            df_tienda['precio_ok'] = None
            if mask.any():
                df_tienda.loc[mask, 'precio_ok'] = abs(df_tienda.loc[mask, 'variacion_precio_%']) <= price_threshold
            
            # Cuotas OK solo si ambas existen
            if selected_store in ["Fravega", "Megatone"] and 'cuotas_maestro' in df_tienda.columns:
                mask_c = ((df_tienda['cuotas'].notna()) & (df_tienda['cuotas_maestro'].notna()) & 
                          (df_tienda['estado_producto'] == 'Activo'))
                df_tienda['cuotas_correctas'] = None
                if mask_c.any():
                    df_tienda.loc[mask_c, 'cuotas_correctas'] = (df_tienda.loc[mask_c, 'cuotas'] == df_tienda.loc[mask_c, 'cuotas_maestro'])
            else:
                df_tienda['cuotas_correctas'] = None
            
            st.session_state.audit_results = df_tienda
            
            st.success(f"✅ Completado: {len(df_tienda)} productos")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("✅ Precio OK", len(df_tienda[df_tienda['precio_ok'] == True]))
            col2.metric("❌ Error precio", len(df_tienda[df_tienda['precio_ok'] == False]))
            col3.metric("⚠️ Inhabilitados", len(df_tienda[df_tienda['estado_producto'] == 'Inhabilitado']))
            col4.metric("🔴 Errores", len(df_tienda[df_tienda['estado_producto'] == 'Error']))

with tab2:
    if st.session_state.audit_results is not None:
        df_results = st.session_state.audit_results
        
        st.markdown("### 📊 Resultados")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            filtros = ["Todos", "Solo activos", "Errores precio", "Inhabilitados", "Errores técnicos"]
            if selected_store in ["Fravega", "Megatone"]:
                filtros.append("Cuotas incorrectas")
            filtro = st.selectbox("Filtrar:", filtros)
        
        df_mostrar = df_results.copy()
        
        if filtro == "Solo activos":
            df_mostrar = df_mostrar[df_mostrar['estado_producto'] == 'Activo']
        elif filtro == "Errores precio":
            df_mostrar = df_mostrar[(df_mostrar['precio_ok'] == False) & (df_mostrar['estado_producto'] == 'Activo')]
        elif filtro == "Inhabilitados":
            df_mostrar = df_mostrar[df_mostrar['estado_producto'] == 'Inhabilitado']
        elif filtro == "Errores técnicos":
            df_mostrar = df_mostrar[df_mostrar['estado_producto'] == 'Error']
        elif filtro == "Cuotas incorrectas":
            df_mostrar = df_mostrar[df_mostrar['cuotas_correctas'] == False]
        
        columnas_mostrar = ['sku', 'titulo', 'precio_maestro', 'precio_web', 'precio_tachado',
                           'descuento_%', 'variacion_precio_%', 'precio_ok', 'categoria', 'estado_producto', 'estado_scraping']
        
        if selected_store in ["Fravega", "Megatone"]:
            columnas_mostrar.insert(8, 'cuotas_maestro')
            columnas_mostrar.insert(9, 'cuotas')
            columnas_mostrar.insert(10, 'cuotas_correctas')
        
        columnas_existentes = [col for col in columnas_mostrar if col in df_mostrar.columns]
        df_display = df_mostrar[columnas_existentes].copy()
        
        # SIN GUIONES BAJOS
        nombres = {
            'sku': 'SKU', 'titulo': 'Título', 'precio_maestro': 'Precio Maestro',
            'precio_web': 'Precio Web', 'precio_tachado': 'Precio Tachado',
            'descuento_%': 'Descuento %', 'variacion_precio_%': 'Variación %',
            'precio_ok': 'Precio OK', 'cuotas_maestro': 'Cuotas Maestro',
            'cuotas': 'Cuotas Web', 'cuotas_correctas': 'Cuotas OK',
            'categoria': 'Categoría', 'estado_producto': 'Estado',
            'estado_scraping': 'Scraping'
        }
        
        df_display = df_display.rename(columns=nombres)
        
        if 'Precio OK' in df_display.columns:
            df_display['Precio OK'] = df_display['Precio OK'].map({True: '✅', False: '❌', None: '-'})
        
        if 'Cuotas OK' in df_display.columns:
            df_display['Cuotas OK'] = df_display['Cuotas OK'].map({True: '✅', False: '❌', None: '-'})
        
        st.dataframe(df_display, use_container_width=True, height=500)
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        with col1:
            excel_file = crear_excel_formateado(df_results, selected_store)
            st.download_button(
                "📊 Descargar Excel",
                data=excel_file,
                file_name=f"Auditoria_{selected_store}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            csv = df_mostrar.to_csv(index=False)
            st.download_button(
                "📄 Descargar CSV",
                data=csv,
                file_name=f"Auditoria_{selected_store}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True
            )
    else:
        st.info("Ejecuta una auditoría primero")

with tab3:
    if st.session_state.audit_results is not None:
        df = st.session_state.audit_results
        
        st.markdown("### 📈 Dashboard")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total = len(df)
            st.metric("📦 Total", total)
        
        with col2:
            activos = len(df[df['estado_producto'] == 'Activo'])
            activos_pct = (activos / total * 100) if total > 0 else 0
            st.metric("✅ Activos", f"{activos} ({activos_pct:.1f}%)")
        
        with col3:
            inhabilitados = len(df[df['estado_producto'] == 'Inhabilitado'])
            st.metric("⚠️ Inhabilitados", inhabilitados)
        
        with col4:
            errores = len(df[df['estado_producto'] == 'Error'])
            st.metric("🔴 Errores", errores)
        
        st.markdown("---")
        col1, col2 = st.columns(2)
        
        with col1:
            estados_data = {
                'Estado': ['Activos', 'Inhabilitados', 'Errores'],
                'Cantidad': [
                    len(df[df['estado_producto'] == 'Activo']),
                    len(df[df['estado_producto'] == 'Inhabilitado']),
                    len(df[df['estado_producto'] == 'Error'])
                ]
            }
            
            fig = px.pie(estados_data, values='Cantidad', names='Estado', title='Distribución de Estados')
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            df_activos = df[df['estado_producto'] == 'Activo']
            if not df_activos.empty and 'precio_ok' in df_activos.columns:
                precios_data = {
                    'Estado': ['✅ Precio OK', '❌ Precio Error'],
                    'Cantidad': [
                        len(df_activos[df_activos['precio_ok'] == True]),
                        len(df_activos[df_activos['precio_ok'] == False])
                    ]
                }
                
                fig = px.pie(precios_data, values='Cantidad', names='Estado', title='Validación de Precios')
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Sin datos de precios")
        
        if selected_store in ["Fravega", "Megatone"] and 'cuotas' in df.columns:
            st.markdown("---")
            st.markdown("### 💳 Análisis de Cuotas")
            
            col1, col2 = st.columns(2)
            
            with col1:
                df_cuotas = df[(df['cuotas'].notna()) & (df['estado_producto'] == 'Activo')]
                if not df_cuotas.empty:
                    cuotas_count = df_cuotas['cuotas'].value_counts().sort_index()
                    fig = px.bar(x=cuotas_count.index, y=cuotas_count.values,
                               title='Distribución de Cuotas', labels={'x': 'Cuotas', 'y': 'Cantidad'})
                    st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                df_cuotas_val = df[(df['cuotas_correctas'].notna()) & (df['estado_producto'] == 'Activo')]
                if not df_cuotas_val.empty:
                    cuotas_ok = len(df_cuotas_val[df_cuotas_val['cuotas_correctas'] == True])
                    cuotas_error = len(df_cuotas_val[df_cuotas_val['cuotas_correctas'] == False])
                    
                    fig = px.pie(values=[cuotas_ok, cuotas_error], names=['✅ Correctas', '❌ Incorrectas'],
                               title='Validación de Cuotas')
                    st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Ejecuta una auditoría primero")

st.markdown("---")
st.markdown(
    f"""<div style='text-align: center; color: gray;'>
        v6.0 FIXED | {datetime.now().strftime("%d/%m/%Y %H:%M")}
    </div>""",
    unsafe_allow_html=True
)
